# coding:utf-8
'''
--------------------------------------------------
[OverView]
Mail:Fileリストを作成します３
[Function]
フォルダ内の/eml拡張子のファイルを探し出してリストアップしていく
[Histry]
2019.01.02 Ver1.0.0 TSJ ogawa
2020.02.03 Ver1.0.1 TSJ　ogawa 
全バックアップなので、ファイルが多くなると遅くなる。差分バックアップにする必要あり

--------------------------------------------------
''' 
import openpyxl
from openpyxl import load_workbook
import os
from email import utils
import time
import shutil #file move 191231
import quopri
#import base64
import io
import sys
import email
from email.header import decode_header

class ExcelOutList(object):
        def __init__(self):
            self.FileCount=0
            MainPath=os.getcwd()#Mainpath
            #ステップ2｜所定フォルダ内の「Book1.xlsm」を指定して読み込む
            Mdirname = os.path.basename(MainPath)
            Mdirname=Mdirname+'_File-list.xlsx'
            self.filepath = os.path.join(MainPath,Mdirname)
            if not os.path.isfile(self.filepath):
                New_wb = openpyxl.Workbook()
                sheet = New_wb.active
                sheet.title = 'List'
                New_wb.save(self.filepath)
                #glob.glob("*.xlsx")
            self.wb = load_workbook(filename=self.filepath)
            self.ws1 = self.wb['List']
            #ステップ3｜集計範囲の取得
            #self.startdate=self.ws1['B2'].value
            #self.enddate=self.ws1['B3'].value
            #Step4|メールファイルからリストの項目となる部分を取り出す。
            for self.folder, self.subfolders, self.files in os.walk(os.getcwd()):
                for file in os.listdir(self.folder):
                    self.base, self.ext = os.path.splitext(file)
                    if self.ext == '.eml':                   
                        self.Bname=os.path.join(self.folder, file)
                        with open(self.Bname, 'rb') as email_file:
                            self.email_message = email.message_from_bytes(email_file.read())
                        self.subject = None
                        self.to_address = None
                        self.cc_address = None
                        self.from_address = None
                        self.date=None #20191224 add date
                        self.body = ""            
                        #print('file:{},ext:{}'.format(file,ext))
                        self.attach_file_list = []
                        #最初に移動してそこのフォルダから処理すれば無限ループに入らないはず
                    # emlの解釈
                        self._makelist(self.folder)
                        #newfilepath = os.path.join(os.getcwd(),'List.xlsx')
                        self.wb.save(self.filepath)
                        print('Seve OK')
                        #print(self.get_format_date(self.date))
        def _makelist(self,Flie0):
            self.date=self._get_decoded_header("Date")# add date
            self.subject = self._get_decoded_header("Subject")
            self.to_address = self._get_decoded_header("To")
            self.cc_address = self._get_decoded_header("Cc")
            self.from_address = self._get_decoded_header("From")
            # メッセージ本文部分の処理
            Attach_count=0 #2019.01.02
            for part in self.email_message.walk():
                # ContentTypeがmultipartの場合は実際のコンテンツはさらに
                # 中のpartにあるので読み飛ばす  タイプが宣言されているのは複数あるので
                if part.get_content_maintype() == 'multipart':
                    continue
                # ファイル名の取得
                attach_fname = part.get_filename()


                # ファイル名がない場合は本文のはず
                if not attach_fname:
                    continue
                else:
                    #2020/01/20
                    if part.get_content_maintype()=="image":
                        print("image")
                        A1 = part.get_all("Content-ID")
                        B1=A1[0].replace("<","")
                        attach_fname = B1.replace(">","_") + attach_fname

                    if part.get_content_maintype()=="application":
                        A1 = decode_header(attach_fname)[0][0]
                        C1 = decode_header(attach_fname)[0][1]
                        attach_fname=A1
                        if not C1:
                            attach_fname=A1
                        else:
                            attach_fname=A1.decode(C1) 
                    self.base1, self.ext1=os.path.splitext(attach_fname)
                    self.base2, self.ext2=self.from_address.split('<')
                    Attach_count=Attach_count+1
                    try:
                        #a=0
                        #hiduketime = datetime.datetime(RT.year ,RT.month, RT.day, RT.hour, RT.minute, RT.second)
                        #if startdate <= hiduketime <= enddate:
                        self.ws1.cell(row=7+self.FileCount, column=1).value = self.FileCount+1   #List counter
                        self.ws1.cell(row=7+self.FileCount, column=2).value = str(attach_fname)  #attach count(添付ファイル名)
                        self.ws1.cell(row=7+self.FileCount, column=3).value = str(self.ext1)     #  ファイル種類
                        self.ws1.cell(row=7+self.FileCount, column=4).value = self.subject       #Subject(件名)
                        self.ws1.cell(row=7+self.FileCount, column=5).value = str(Flie0)         #File name  
                        self.ws1.cell(row=7+self.FileCount, column=6).value = self.base2         #From(送信者)
                        self.ws1.cell(row=7+self.FileCount, column=7).value = self.to_address    #To(送信先)
                        self.ws1.cell(row=7+self.FileCount, column=8).value = self.get_format_date(self.date)#self.get_format_date(self.date)         #data(送受信日付)

                        self.FileCount=self.FileCount+1
                    except:
                        print('cannot save ' + attach_fname)    
            return 1
        def _get_decoded_header(self, key_name):
            """
            ヘッダーオブジェクトからデコード済の結果を取得する
            """
            ret = ""

            # 該当項目がないkeyは空文字を戻す
            raw_obj = self.email_message.get(key_name)
            if raw_obj is None:
                return ""
            # デコードした結果をunicodeにする
            for fragment, encoding in decode_header(raw_obj):
                if not hasattr(fragment, "decode"):
                    ret += fragment
                    continue
                # encodeがなければとりあえずUTF-8でデコードする
                if encoding:
                    ret += fragment.decode(encoding)
                else:
                    ret += fragment.decode("UTF-8")
            return ret  
        def get_format_date(self, date_string):
            """
            メールの日付をtimeに変換
            http://www.faqs.org/rfcs/rfc2822.html
            "Jan" / "Feb" / "Mar" / "Apr" /"May" / "Jun" / "Jul" / "Aug" /"Sep" / "Oct" / "Nov" / "Dec"
            Wed, 12 Dec 2007 19:18:10 +0900
            """
            format_pattern = '%Y%m%d %H%M%S' #'%a, %d %b %Y %H:%M:%S'

            #3 Jan 2012 17:58:09という形式でくるパターンもあるので、
            #先頭が数値だったらパターンを変更
            if date_string[0].isdigit():
                format_pattern = '%d %b %Y %H:%M:%S'

            #time_tuple = parsedate_tz(date_string)
            time_tuple=utils.parsedate(date_string)
            return time.strftime('%Y/%m/%d  %H:%M:%S \n',time_tuple )
            # return datetime.strptime(date_string[0:-6],format_pattern)#date_string[0:-6]                   
if __name__ == "__main__":
    result = ExcelOutList()
    print(result)