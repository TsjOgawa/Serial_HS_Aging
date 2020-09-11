# coding:utf-8
'''
--------------------------------------------------
[OverView]
Mail解析スクリプト
[Function]
フォルダ内の/eml拡張子のファイルを探し出して
リストアップしていく
[Histry]
2019.01.02 Ver1.0.0 TSJ ogawa
2019.01.02 Ver1.0.1 TSJ ogawa　添付無でもＯＫなようにした.
2020.01.20 Ver1.0.2 TSJ ogawa
メール本文に張り付けてあるイメージに問題発生。同じ名前で保存するので
最後の画像のみが残ってしまう。修正が必要
本文に貼ってある画像か、添付画像化を区別する必要がある（添付資料としての画像は、ファイル名があるので）
2020.09.08 Ver1.0.3 ogawa
検索機能を簡略化。カレントフォルダ内のすべてのファイルからメールファイルを検索していたが
カレントフォルダに置いたメールだけ検索するように変更。
時間短縮を図る
--------------------------------------------------
''' 
#===LIblaly file===
import openpyxl
from openpyxl import load_workbook
from email import utils
from email.utils import parsedate_tz, mktime_tz
from datetime import datetime 
import time
import shutil #file move 191231
import quopri
import base64
import io
import os
import sys
import email
from email.header import decode_header

#このコードだとフォルダにスクリプトを持っていく必要がある
class MailParser(object):
    """
    メールファイルのパスを受け取り、それを解析するクラス
    """

    def __init__(self):
        #Main pathを記録します
        #------------------------------------------------------------------------
        self.FileCount=0
        MainPath=os.getcwd()#Mainpath
        #ステップ2｜所定フォルダ内の「Book1.xlsm」を指定して読み込む
        Mdirname = os.path.basename(MainPath)
        Mdirname=Mdirname+'_File-list.xlsx'
        self.filepath = os.path.join(MainPath,Mdirname)
        if not os.path.isfile(self.filepath):
            New_wb = openpyxl.Workbook()
            sheet = New_wb.active
            sheet.title = 'List'
            New_wb.save(self.filepath)
            #glob.glob("*.xlsx")
        self.wb = load_workbook(filename=self.filepath,data_only=True)
        #self.wb = load_workbook(filename=self.filepath)
        self.ws1 = self.wb['List']
        #self.ws0 = self.wb0['List']
        #ステップ3｜集計範囲の取得
        self.FileCount=self.ws1['B3'].value
        #self.ws1['B3'].value=str("=COUNTA(A7:A1048576)")
        #self.startdate=self.ws1['B2'].value
        #self.enddate=self.ws1['B3'].value
        #Step4|メールファイルからリストの項目となる部分を取り出す。}
        #------------------------------------------------------------------------
        print(os.getcwd())
        #for self.folder, self.subfolders, self.files in os.walk(os.getcwd()):
            #既にあるフォルダ内は検索しない方向で調整する
        for file in os.listdir(os.getcwd()):
        #for file in os.listdir(self.folder):
            #self.base, self.ext = os.path.splitext(self.folder)#2020/09/07
            self.base, self.ext = os.path.splitext(file)
            if self.ext == '.eml': 
                self.Bname=os.path.join(os.getcwd(), file)
                #self.Bname=os.path.join(self.folder, file)
                with open(self.Bname, 'rb') as email_file:
                    self.email_message = email.message_from_bytes(email_file.read())
                self.subject = None
                self.to_address = None
                self.cc_address = None
                self.from_address = None
                self.date=None #20191224 add date
                self.body = ""            
                #print('file:{},ext:{}'.format(file,ext))
                self.attach_file_list = []
                #最初に移動してそこのフォルダから処理すれば無限ループに入らないはず
                self.date=self._get_decoded_header("Date")# add date
            # emlの解釈
            #ここでファイルが既にあるかをチェックする
            #ファイルがあれば無視して次のファイルを確認する
                print(self.get_format_date(0,self.date))
                print(os.path.join(os.getcwd(),self.get_format_date(0,self.date)+'_'+self.base))
                if not  os.path.isdir(os.path.join(os.getcwd(),self.get_format_date(0,self.date)+'_'+self.base)):
                    self._parse(os.getcwd())#--
                    #self._parse(self.folder)#--
                    print(self.get_attr_data())
                    self.wb.save(self.filepath)
                    print('List Seve OK')
                else:
                    continue
                        
                        #print(self.get_format_date(self.date))
            
            """     if ext == '.eml':
                    print('files: {}'.format(files)) """

            """  
            print('folder: {}'.format(folder))
                print('subfolders: {}'.format(subfolders))
                print('files: {}'.format(files))
                print() 
            """
    def get_attr_data(self):
        try:
            """
            メールデータの取得
            """
            result = """\
        DATE: {}
        FROM: {}
        TO: {}
        CC: {}
        -----------------------
        BODY:
        {}
        -----------------------
        ATTACH_FILE_NAME:
        {}
        """.format(
                self.date,
                self.from_address,
                self.to_address,
                self.cc_address,
                self.body,
                "_AND_".join([ x["name"] for x in self.attach_file_list])
            )
            return result
        except:
            print('cannot saved file ')


    def _parse(self,Flie0):
        self.date=self._get_decoded_header("Date")# add date
        self.subject = self._get_decoded_header("Subject")
        self.to_address = self._get_decoded_header("To")
        self.cc_address = self._get_decoded_header("Cc")
        self.from_address = self._get_decoded_header("From")
        self.NewFile=(self.get_format_date(0,self.date)+'_'+self.base)
        Flie0=os.path.join(os.getcwd(),self.NewFile)
        Attach_count=0 #2019.01.02
        if not os.path.isdir(Flie0):
            os.makedirs(Flie0)
        # メッセージ本文部分の処理
            for part in self.email_message.walk():
                # ContentTypeがmultipartの場合は実際のコンテンツはさらに
                # 中のpartにあるので読み飛ばす  タイプが宣言されているのは複数あるので
                if part.get_content_maintype() == 'multipart':
                    continue
                # ファイル名の取得
                attach_fname = part.get_filename()
                # ファイル名がない場合は本文のはず
                if not attach_fname:
                    if part.get_content_maintype() == 'text':  
                        charset = str(part.get_content_charset())
                        if charset:
                            self.body += part.get_payload(decode=True).decode(charset, errors="replace")
                        else:
                            self.body += part.get_payload(decode=True)
                else:
                    # ファイル名があるならそれは添付ファイルなので
                    # データを取得する
                    '''
                    ---------------------------------------
                    ここから新しいコードを追加します。
                    [Function]
                    --------------------------------------
                    '''
                    if part.get_content_maintype()=="image":
                     #   if attach_fname.find("image")!=-1: #1　コンテンツチェックをやめる
                        A1 = part.get_all("Content-ID")
                        if A1!=None:
                            B1=A1[0].replace("<","")
                            attach_fname = B1.replace(">","_") + attach_fname
                        #A1 = decode_header(attach_fname)[0][0]
                    elif part.get_content_maintype()=="application":
                        A1 = decode_header(attach_fname)[0][0]
                        C1 = decode_header(attach_fname)[0][1]
                        attach_fname=A1
                        if not C1:
                            attach_fname=A1
                        else:
                            attach_fname=A1.decode(C1) 
                    self.base1, self.ext1=os.path.splitext(attach_fname)
                    self.base2, self.ext2=self.from_address.split('<')
                    Attach_count=Attach_count+1        
                    #print("Test_title:"+attach_fname)
                        #attach_fname=A1.decode(C1) 
                
                    try:
                    #    if part.get_content_maintype()=="application":
                    #          attach_fname='test.zip'
                    #  ここで新しいフォルダを作成する
                        #Flie0=os.path.join(os.getcwd(),self.NewFile)
                        #ファイルがすでに存在するかを確認する=>存在しても中身がない場合があるのでVer1.0.0ではそのまま作業を行う
                        
                        if not os.path.isdir(Flie0):
                            os.makedirs(Flie0)
                        #なんども同じ作業を繰り返すことになるが、それは後々更新
                        #既に存在するファイルは繰り返さない（時間がかかるのと差分更新にしたいので）2020.01.09
                        #--------------------------------------------------------------------------------------------------------
                        self.ws1.cell(row=7+self.FileCount, column=1).value = self.FileCount+1   #List counter
                        self.ws1.cell(row=7+self.FileCount, column=2).value = str(attach_fname)  #attach count(添付ファイル名)
                        self.ws1.cell(row=7+self.FileCount, column=3).value = str(self.ext1)     #  ファイル種類
                        self.ws1.cell(row=7+self.FileCount, column=4).value = self.subject       #Subject(件名)
                        self.ws1.cell(row=7+self.FileCount, column=5).value = str(Flie0)         #File name  
                        self.ws1.cell(row=7+self.FileCount, column=6).value = self.base2         #From(送信者)
                        self.ws1.cell(row=7+self.FileCount, column=7).value = self.to_address    #To(送信先)
                        self.ws1.cell(row=7+self.FileCount, column=8).value = self.get_format_date(1,self.date)#self.get_format_date(self.date)         #data(送受信日付)
                        self.ws1.cell(row=7+self.FileCount, column=8).number_format="yyyy/m/d h:mm"

                        self.FileCount=self.FileCount+1
                        self.ws1['B3'].value=int(self.FileCount) 
                        with open(os.path.join(Flie0, attach_fname), 'wb' ) as f:  # 20200908
                            f.write(part.get_payload(None, True)) 
                            '''
                                if part.get_content_maintype()=="application":
                                    print(attach_fname.decode("utf-8")) 
                                    test1=part.get_payload()
                                    test1=base64.urlsafe_b64decode(test1.encode('ASCII')).decode("utf-8")
                                    f.write(test1) 
                                else:
                                    f.write(part.get_payload(decode=True))             # N
                        #     f.write(io.BytesIO(part.get_payload(decode=True)))
                        '''
                    except:
                        print('cannot save ' + str(attach_fname))
                        print(attach_fname)
                    self.attach_file_list.append({
                        "name": attach_fname,
                        "data": part.get_payload(decode=True)
                    })
            #ここで総合的な処理を行う メールファイルを移動
            
            try:
                if not os.path.isdir(Flie0):
                    os.makedirs(Flie0)
                #本文をテキストで保存する
                with open(os.path.join(Flie0, 'メール本文.txt'), 'wb' ) as T:  # M
                    T.write(self.body.encode('utf-8'))

                with open(os.path.join(Flie0, 'メールCC.txt'), 'wb' ) as T:  # M
                    T.write(self.cc_address.encode('utf-8'))

                if os.path.isdir(Flie0):
                    shutil.move(self.Bname,Flie0)
            except:
                print('既にメールファイルは移動しています')

    def _get_decoded_header(self, key_name):
        """
        ヘッダーオブジェクトからデコード済の結果を取得する
        """
        ret = ""

        # 該当項目がないkeyは空文字を戻す
        raw_obj = self.email_message.get(key_name)
        if raw_obj is None:
            return ""
        # デコードした結果をunicodeにする
        for fragment, encoding in decode_header(raw_obj):
            if not hasattr(fragment, "decode"):
                ret += fragment
                continue
            # encodeがなければとりあえずUTF-8でデコードする
            if encoding:
                ret += fragment.decode(encoding)
            else:
                ret += fragment.decode("UTF-8")
        return ret
    def get_format_date(self, Ftype,date_string):
        """
        メールの日付をtimeに変換
        http://www.faqs.org/rfcs/rfc2822.html
        "Jan" / "Feb" / "Mar" / "Apr" /"May" / "Jun" / "Jul" / "Aug" /"Sep" / "Oct" / "Nov" / "Dec"
        Wed, 12 Dec 2007 19:18:10 +0900
        """
        format_pattern = '%Y%m%d %H%M%S' #'%a, %d %b %Y %H:%M:%S'

        #3 Jan 2012 17:58:09という形式でくるパターンもあるので、
        #先頭が数値だったらパターンを変更
        if date_string[0].isdigit():
            format_pattern = '%d %b %Y %H:%M:%S'

        #time_tuple = parsedate_tz(date_string)
        if Ftype==0:
          format_pattern = '[%Y%m%d_%H%M%S]'
        else:
            format_pattern ='%Y/%m/%d %H:%M:%S'
        time_tuple=utils.parsedate(date_string)
        return time.strftime(format_pattern,time_tuple )
        # return datetime.strptime(date_string[0:-6],format_pattern)#date_string[0:-6]
    def my_makedirs(self,path):
        '''
        --------------------------------------------
        フォルダが存在するかどうかを確認して
        存在すれば、そのまま終了
    　　　存在がなければ新たに作成してフォルダ名を変更
    　　---------------------------------------------
        '''
        if not os.path.isdir(path):
            os.makedirs(path)
        return 1


if __name__ == "__main__":
    result = MailParser().get_attr_data()
    #result = MailParser()
    print(result)